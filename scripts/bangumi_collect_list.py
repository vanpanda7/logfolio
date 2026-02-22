#!/usr/bin/env python3
"""
从 Bangumi 官方 API 拉取「看过的动画」并保存为 Excel。
文档: https://bangumi.github.io/api/
依赖: pip install openpyxl
"""
import argparse
import json
import os
import time
import urllib.request
import urllib.parse

API_BASE = "https://api.bgm.tv"
HEADERS = {
    "User-Agent": "Logfolio/1.0 (https://github.com/your-repo; bangumi collect list)",
    "Accept": "application/json",
}


def fetch_collections(username: str, subject_type: int = 2, collect_type: int = 2, token: str | None = None):
    """subject_type: 2=动画, type: 2=看过. 私有收藏需传 token."""
    all_items = []
    offset = 0
    limit = 50
    while True:
        params = {
            "subject_type": subject_type,
            "type": collect_type,
            "limit": limit,
            "offset": offset,
        }
        url = f"{API_BASE}/v0/users/{username}/collections?{urllib.parse.urlencode(params)}"
        req = urllib.request.Request(url, headers=dict(HEADERS))
        if token:
            req.add_header("Authorization", f"Bearer {token}")
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.loads(r.read().decode())
        items = data.get("data") or []
        if not items:
            break
        for item in items:
            sub = item.get("subject") or {}
            name_cn = (sub.get("name_cn") or "").strip()
            name = (sub.get("name") or "").strip()
            title = name_cn or name or ""
            if title:
                all_items.append({
                    "title": title,
                    "id": sub.get("id"),
                    "name_jp": name if name_cn else (sub.get("name") or ""),
                    "subject": sub,
                    "raw_item": item,
                })
        if len(items) < limit:
            break
        offset += limit
        time.sleep(0.3)
    return all_items


def save_excel(items: list, path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise SystemExit("请先安装: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "看过的动画"

    headers = ["序号", "标题", "日文名", "条目ID", "我的评分", "收藏更新时间", "Bangumi链接"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", wrap_text=True)

    for i, it in enumerate(items, 1):
        raw = it.get("raw_item") or {}
        sub = it.get("subject") or {}
        updated = raw.get("updated_at") or ""
        if isinstance(updated, str) and "T" in updated:
            updated = updated.replace("T", " ").split(".")[0]
        score = raw.get("rate") or sub.get("rating", {}).get("score") or ""
        sid = sub.get("id") or ""
        link = f"https://bgm.tv/subject/{sid}" if sid else ""
        ws.append([
            i,
            it["title"] or "",
            it.get("name_jp") or sub.get("name") or "",
            sid,
            score,
            updated,
            link,
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 28

    wb.save(path)
    print(f"已保存到: {os.path.abspath(path)}（共 {len(items)} 条）")


def main():
    parser = argparse.ArgumentParser(description="拉取 Bangumi 用户「看过的动画」并保存为 Excel")
    parser.add_argument("username", nargs="?", default="1216434", help="Bangumi 用户名或 ID（默认 1216434）")
    parser.add_argument("--token", "-t", help="Access Token（仅查看私有收藏时需要）")
    parser.add_argument("-o", "--output", default="bangumi_看过_动画.xlsx", help="输出 Excel 文件路径（默认 bangumi_看过_动画.xlsx）")
    parser.add_argument("--json", "-j", action="store_true", help="额外输出完整 JSON 到同目录")
    parser.add_argument("--titles-only", action="store_true", help="仅打印标题到控制台，不写 Excel")
    args = parser.parse_args()

    items = fetch_collections(args.username, subject_type=2, collect_type=2, token=args.token)
    if not items:
        print("未获取到任何条目")
        return

    if args.titles_only:
        for it in items:
            print(it["title"])
        return

    save_excel(items, args.output)
    if args.json:
        json_path = os.path.splitext(args.output)[0] + ".json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump([{"title": it["title"], "id": it["id"], "subject": it.get("subject")} for it in items], f, ensure_ascii=False, indent=2)
        print(f"JSON 已保存: {json_path}")


if __name__ == "__main__":
    main()
