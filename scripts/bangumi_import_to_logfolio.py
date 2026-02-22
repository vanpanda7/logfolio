#!/usr/bin/env python3
"""
从 bangumi_看过_动画.xlsx 批量导入到 Logfolio：每条创建为「历史」记录，并拉取 Bangumi 封面。

支持两种 Excel 格式：
  - 仅「标题」一列：按标题在 Bangumi 搜索动画，取第一个结果作为封面；
  - 含「序号/标题/日文名/条目ID」：直接用条目ID 拉取封面（更准）。

需先在 Logfolio 建好「动漫」分类。请求会带 X-User-ID，与前端登录用户一致。
"""
import argparse
import json
import os
import time
import urllib.request
import urllib.parse

BGM_API = "https://api.bgm.tv"
BGM_HEADERS = {
    "User-Agent": "Logfolio/1.0 (bangumi import)",
    "Accept": "application/json",
}


def read_excel(path: str):
    """读取 Excel，返回 [(标题, 条目ID或None), ...]。若表头含「条目ID」则用其取封面，否则仅标题（后面用搜索取封面）。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("请先安装: pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    first = rows[0]
    # 判断表头：有「条目ID」或第4列为数字则认为是 序号,标题,日文名,条目ID
    has_id_col = False
    if len(first) >= 4:
        fourth = first[3]
        if fourth and "条目" in str(fourth) or "ID" in str(fourth):
            has_id_col = True
    out = []
    for row in rows[1:]:
        if not row:
            continue
        title = (row[1] if len(row) > 1 else row[0] if row else "").strip()
        if not title:
            title = (row[0] or "").strip()
        if not title:
            continue
        sid = None
        if has_id_col and len(row) > 3:
            sid = row[3]
            if isinstance(sid, (int, float)):
                sid = int(sid)
            else:
                try:
                    sid = int(str(sid).strip()) if sid else None
                except (ValueError, TypeError):
                    sid = None
        out.append((title, sid))
    return out


def search_bangumi_subject(keyword: str) -> dict | None:
    """按关键词搜索动画，返回第一个结果的 subject 信息（含 id、images）。"""
    url = f"{BGM_API}/v0/search/subjects"
    body = json.dumps({"keyword": keyword, "filter": {"type": [2]}}).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST", headers={
        **BGM_HEADERS,
        "Content-Type": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read().decode())
        items = data.get("data") or []
        if not items:
            return None
        return items[0]
    except Exception:
        return None


def get_bangumi_cover_url(subject_id: int) -> str | None:
    """获取 Bangumi 条目封面大图 URL（lain.bgm.tv，Logfolio 后端允许）"""
    url = f"{BGM_API}/v0/subjects/{subject_id}"
    req = urllib.request.Request(url, headers=dict(BGM_HEADERS))
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read().decode())
        images = data.get("images") or {}
        return (images.get("large") or images.get("medium") or images.get("common") or "").strip() or None
    except Exception:
        return None


def get_categories(api_base: str, user_id: str, auth: tuple | None) -> list:
    """获取分类列表"""
    url = api_base.rstrip("/") + "/categories/"
    req = urllib.request.Request(url, headers={"Accept": "application/json", "X-User-ID": user_id})
    if auth:
        import base64
        req.add_header("Authorization", "Basic " + base64.b64encode(f"{auth[0]}:{auth[1]}".encode()).decode())
    with urllib.request.urlopen(req, timeout=15) as r:
        data = json.loads(r.read().decode())
    return data if isinstance(data, list) else []


def create_item(api_base: str, user_id: str, auth: tuple | None, category_id: int, title: str, cover_url: str | None) -> dict | None:
    """创建一条历史记录（skip_finish_time=1），带封面"""
    url = api_base.rstrip("/") + "/items/"
    form = {
        "title": title,
        "category_id": category_id,
        "is_completed": "true",
        "skip_finish_time": "1",
    }
    if cover_url:
        form["cover_image_url"] = cover_url
    body = urllib.parse.urlencode(form).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST", headers={
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
        "X-User-ID": user_id,
    })
    if auth:
        import base64
        req.add_header("Authorization", "Basic " + base64.b64encode(f"{auth[0]}:{auth[1]}".encode()).decode())
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        try:
            err = e.read().decode()
        except Exception:
            err = str(e)
        raise SystemExit(f"创建失败 [{e.code}]: {err}")
    except Exception as e:
        raise SystemExit(f"请求失败: {e}")


def main():
    parser = argparse.ArgumentParser(description="从 Bangumi 看过列表 Excel 批量导入 Logfolio 历史记录并拉取封面")
    parser.add_argument("excel", nargs="?", default="bangumi_看过_动画.xlsx", help="Excel 文件路径")
    parser.add_argument("--api", "-a", default=os.environ.get("LOGFOLIO_API", "http://127.0.0.1:8000/api"), help="Logfolio API 根地址")
    parser.add_argument("--user-id", "-u", default=os.environ.get("LOGFOLIO_USER_ID", "rick"), help="X-User-ID，与前端登录用户一致")
    parser.add_argument("--user", help="Basic 认证用户名（若 API 前有鉴权）")
    parser.add_argument("--password", help="Basic 认证密码")
    parser.add_argument("--dry-run", action="store_true", help="只列出将要导入的条目，不请求 API")
    parser.add_argument("--limit", "-n", type=int, default=0, help="最多导入条数，0 表示全部")
    parser.add_argument("--skip-cover", action="store_true", help="不拉取封面，仅创建标题")
    args = parser.parse_args()

    excel_path = os.path.abspath(args.excel)
    if not os.path.isfile(excel_path):
        raise SystemExit(f"文件不存在: {excel_path}")

    rows = read_excel(excel_path)
    if not rows:
        raise SystemExit("Excel 中未读到有效行（需有 标题 与 条目ID 列）")

    if args.limit:
        rows = rows[: args.limit]
    print(f"共 {len(rows)} 条待导入（Excel: {excel_path}）")

    auth = (args.user, args.password) if args.user and args.password else None
    api_base = args.api.rstrip("/")
    if not api_base.endswith("/api"):
        api_base = api_base + "/api"

    categories = get_categories(api_base, args.user_id, auth)
    anime_cat = next((c for c in categories if (c.get("name") or "").strip() == "动漫"), None)
    if not anime_cat:
        names = [c.get("name") for c in categories]
        raise SystemExit(f"未找到「动漫」分类，当前分类: {names}。请先在 Logfolio 创建「动漫」分类。")
    category_id = anime_cat["id"]
    print(f"使用分类: 动漫 (id={category_id})")

    if args.dry_run:
        for i, (title, sid) in enumerate(rows, 1):
            print(f"  {i}. {title} (subject_id={sid})")
        return

    ok, fail = 0, 0
    for i, (title, sid) in enumerate(rows, 1):
        cover_url = None
        if not args.skip_cover:
            if sid:
                cover_url = get_bangumi_cover_url(sid)
                time.sleep(0.2)
            else:
                sub = search_bangumi_subject(title)
                if sub:
                    imgs = sub.get("images") or {}
                    cover_url = (imgs.get("large") or imgs.get("medium") or imgs.get("common") or "").strip() or None
                time.sleep(0.4)
        try:
            create_item(api_base, args.user_id, auth, category_id, title, cover_url)
            ok += 1
            print(f"[{i}/{len(rows)}] 已创建: {title[:36]}{'…' if len(title) > 36 else ''}" + (" (含封面)" if cover_url else ""))
        except Exception as e:
            fail += 1
            print(f"[{i}/{len(rows)}] 失败: {title[:36]} — {e}")
        time.sleep(0.3)
    print(f"\n完成: 成功 {ok}, 失败 {fail}")


if __name__ == "__main__":
    main()
